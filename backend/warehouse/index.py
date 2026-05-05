"""
Склад: управление остатками товаров, документы прихода и списания.
GET  /?action=list          — список товаров с остатками
GET  /?action=movements     — история движений (фильтр: product_id, type, date_from, date_to)
POST / {action:income}      — приход товара (document_number, items: [{product_id, amount_ml}], comment)
POST / {action:writeoff}    — списание по акту (document_number, items: [{product_id, amount_ml}], comment)
POST / {action:export}      — экспорт книги движений в Excel (date_from, date_to)
"""
import json
import os
import io
import psycopg2


def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'])


CORS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type, X-Auth-Token',
}


def get_admin_user(headers: dict):
    token = (headers.get('X-Auth-Token') or '').strip()
    if not token:
        return None
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT u.id, u.role FROM sessions s JOIN users u ON s.user_id = u.id WHERE s.id = %s AND s.expires_at > NOW()",
        (token,)
    )
    row = cur.fetchone()
    conn.close()
    if not row or row[1] not in ('admin', 'moderator'):
        return None
    return {'id': row[0], 'role': row[1]}


def handler(event: dict, context) -> dict:
    """Управление складом: приход, списание, остатки, экспорт."""
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': CORS, 'body': ''}

    method = event.get('httpMethod', 'GET')
    headers = event.get('headers') or {}
    params = event.get('queryStringParameters') or {}

    user = get_admin_user(headers)
    if not user:
        return {'statusCode': 403, 'headers': CORS, 'body': json.dumps({'error': 'Доступ запрещён'})}

    if method == 'GET':
        action = params.get('action', 'list')

        if action == 'list':
            conn = get_conn()
            cur = conn.cursor()
            name_filter = (params.get('name') or '').strip().lower()
            query = """
                SELECT p.id, p.name, p.brand, p.price_per_ml, p.stock_ml, p.booked_ml, p.is_active
                FROM products p
                WHERE 1=1
            """
            values = []
            if name_filter:
                query += " AND (LOWER(p.name) LIKE %s OR LOWER(p.brand) LIKE %s)"
                values.extend([f'%{name_filter}%', f'%{name_filter}%'])
            query += " ORDER BY p.brand, p.name"
            cur.execute(query, values)
            rows = cur.fetchall()
            conn.close()
            products = [{
                'id': r[0], 'name': r[1], 'brand': r[2],
                'price_per_ml': float(r[3]),
                'stock_ml': float(r[4]),
                'booked_ml': float(r[5]),
                'available_ml': float(r[4]) - float(r[5]),
                'is_active': r[6],
            } for r in rows]
            return {'statusCode': 200, 'headers': CORS, 'body': json.dumps({'products': products})}

        if action == 'movements':
            product_id = params.get('product_id')
            mov_type = params.get('type')
            date_from = params.get('date_from')
            date_to = params.get('date_to')

            conn = get_conn()
            cur = conn.cursor()
            query = """
                SELECT sm.id, sm.product_id, p.name, p.brand,
                       sm.type, sm.amount_ml, sm.document_number,
                       sm.order_id, sm.comment, sm.created_at,
                       u.nickname
                FROM stock_movements sm
                JOIN products p ON sm.product_id = p.id
                LEFT JOIN users u ON sm.created_by = u.id
                WHERE 1=1
            """
            values = []
            if product_id:
                query += " AND sm.product_id = %s"
                values.append(int(product_id))
            if mov_type:
                query += " AND sm.type = %s"
                values.append(mov_type)
            if date_from:
                query += " AND sm.created_at >= %s"
                values.append(date_from)
            if date_to:
                query += " AND sm.created_at < (%s::date + INTERVAL '1 day')"
                values.append(date_to)
            query += " ORDER BY sm.created_at DESC LIMIT 500"
            cur.execute(query, values)
            rows = cur.fetchall()
            conn.close()
            movements = [{
                'id': r[0], 'product_id': r[1], 'product_name': r[2], 'brand': r[3],
                'type': r[4], 'amount_ml': float(r[5]), 'document_number': r[6],
                'order_id': r[7], 'comment': r[8],
                'created_at': str(r[9]), 'created_by': r[10],
            } for r in rows]
            return {'statusCode': 200, 'headers': CORS, 'body': json.dumps({'movements': movements})}

        return {'statusCode': 404, 'headers': CORS, 'body': json.dumps({'error': 'Not found'})}

    if method == 'POST':
        body = json.loads(event.get('body') or '{}')
        action = body.get('action', '')

        if action == 'income':
            document_number = (body.get('document_number') or '').strip()
            comment = (body.get('comment') or '').strip()
            items = body.get('items', [])
            if not items:
                return {'statusCode': 400, 'headers': CORS, 'body': json.dumps({'error': 'Укажите items'})}

            conn = get_conn()
            cur = conn.cursor()
            added = []
            for item in items:
                pid = int(item['product_id'])
                amount = float(item['amount_ml'])
                if amount <= 0:
                    continue
                cur.execute(
                    "UPDATE products SET stock_ml = stock_ml + %s WHERE id = %s",
                    (amount, pid)
                )
                cur.execute(
                    """INSERT INTO stock_movements (product_id, type, amount_ml, document_number, comment, created_by)
                       VALUES (%s, 'income', %s, %s, %s, %s) RETURNING id""",
                    (pid, amount, document_number or None, comment or None, user['id'])
                )
                added.append({'movement_id': cur.fetchone()[0], 'product_id': pid, 'amount_ml': amount})
            conn.commit()
            conn.close()
            return {'statusCode': 200, 'headers': CORS, 'body': json.dumps({'ok': True, 'added': added})}

        if action == 'writeoff':
            document_number = (body.get('document_number') or '').strip()
            comment = (body.get('comment') or '').strip()
            items = body.get('items', [])
            if not items:
                return {'statusCode': 400, 'headers': CORS, 'body': json.dumps({'error': 'Укажите items'})}

            conn = get_conn()
            cur = conn.cursor()
            written = []
            for item in items:
                pid = int(item['product_id'])
                amount = float(item['amount_ml'])
                if amount <= 0:
                    continue
                cur.execute(
                    "UPDATE products SET stock_ml = GREATEST(0, stock_ml - %s) WHERE id = %s",
                    (amount, pid)
                )
                cur.execute(
                    """INSERT INTO stock_movements (product_id, type, amount_ml, document_number, comment, created_by)
                       VALUES (%s, 'writeoff', %s, %s, %s, %s) RETURNING id""",
                    (pid, amount, document_number or None, comment or None, user['id'])
                )
                written.append({'movement_id': cur.fetchone()[0], 'product_id': pid, 'amount_ml': amount})
            conn.commit()
            conn.close()
            return {'statusCode': 200, 'headers': CORS, 'body': json.dumps({'ok': True, 'written': written})}

        if action == 'export':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            import base64

            date_from = body.get('date_from')
            date_to = body.get('date_to')

            conn = get_conn()
            cur = conn.cursor()
            query = """
                SELECT sm.created_at, sm.type, p.brand, p.name,
                       sm.amount_ml, sm.document_number, sm.comment,
                       sm.order_id, u.nickname
                FROM stock_movements sm
                JOIN products p ON sm.product_id = p.id
                LEFT JOIN users u ON sm.created_by = u.id
                WHERE 1=1
            """
            values = []
            if date_from:
                query += " AND sm.created_at >= %s"
                values.append(date_from)
            if date_to:
                query += " AND sm.created_at < (%s::date + INTERVAL '1 day')"
                values.append(date_to)
            query += " ORDER BY sm.created_at ASC"
            cur.execute(query, values)
            rows = cur.fetchall()
            conn.close()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Движение товаров'

            header_font = Font(bold=True, color='FFFFFF')
            income_fill = PatternFill('solid', fgColor='1A7F37')
            writeoff_fill = PatternFill('solid', fgColor='CF2424')
            order_fill = PatternFill('solid', fgColor='9B3BCC')
            header_fill = PatternFill('solid', fgColor='1E293B')

            headers_row = ['Дата', 'Тип', 'Бренд', 'Товар', 'Количество (мл)', 'Документ / Акт', 'Комментарий', '№ Заказа', 'Кто']
            ws.append(headers_row)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            TYPE_LABELS = {
                'income': 'Приход',
                'writeoff': 'Списание',
                'order_writeoff': 'Отгрузка покупателю',
            }

            for r in rows:
                created_at, mov_type, brand, name, amount_ml, doc_num, comment, order_id, nickname = r
                label = TYPE_LABELS.get(mov_type, mov_type)
                doc_str = doc_num or (f'Заказ #{order_id}' if order_id else '')
                ws.append([
                    created_at.strftime('%d.%m.%Y %H:%M') if created_at else '',
                    label,
                    brand, name,
                    float(amount_ml),
                    doc_str,
                    comment or '',
                    order_id or '',
                    nickname or '',
                ])
                last_row = ws.max_row
                fill = income_fill if mov_type == 'income' else (order_fill if mov_type == 'order_writeoff' else writeoff_fill)
                ws.cell(last_row, 2).fill = fill
                ws.cell(last_row, 2).font = Font(color='FFFFFF', bold=True)

            col_widths = [18, 22, 20, 30, 16, 22, 30, 10, 15]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = w

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            xlsx_b64 = base64.b64encode(buf.read()).decode()

            return {
                'statusCode': 200,
                'headers': {**CORS, 'Content-Type': 'application/json'},
                'body': json.dumps({'ok': True, 'xlsx_base64': xlsx_b64, 'filename': 'warehouse_movements.xlsx'}),
            }

        return {'statusCode': 404, 'headers': CORS, 'body': json.dumps({'error': 'Not found'})}

    return {'statusCode': 405, 'headers': CORS, 'body': json.dumps({'error': 'Method not allowed'})}
